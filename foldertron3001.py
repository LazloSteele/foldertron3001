from openpyxl import *
import ft_gui as b
import os

def make_folder(folder_name, directory):
    path = os.path.join(directory, folder_name)

    if not os.path.exists(path):
        os.mkdir(path)

def parse_excel(excel, directory):
    wb = load_workbook(excel)
    ws = wb.active

    row_iter = 1
    for row in ws['A']:
        if row.value != None and row.value != '#' and row.value != ' ':
            evt_number = str(row.value).zfill(3)
            evt_level = ws[f"C{row_iter}"].value.replace("/","&")
            evt_title = ws[f"D{row_iter}"].value.replace("/","&")
            make_folder(f"{evt_number} {evt_title}", directory)
        row_iter += 1

def main():
    window = b.browser()



if __name__ == "__main__":
    main()
